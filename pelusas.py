
import requests
import json
import random
import time
import re
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from rich.console import Console
from datetime import datetime

console = Console() # Para logging general del servidor y IAs

# Lista inicial de modelos de IA con claves API
MODELS = [
    {
        "name": "Google Gemini 2.0 Flash",
        "id": "gemini-2.0-flash",
        "api_url": "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent",
        "api_key": "AIzaSyCoFRhRa1zamcMIqIxIuBChtehsoRE5AUM", # Reemplazar con tu clave real o variable de entorno
        "provider": "google"
    },
    {
        "name": "DeepSeek Chat",
        "id": "deepseek-chat",
        "api_url": "https://api.deepseek.com/chat/completions",
        "api_key": "sk-63de741c3c104643816a2c1ba585c63c", # Reemplazar
        "provider": "deepseek"
    },
    {
        "name": "Google Gemini 2.5 Flash",
        "id": "gemini-2.5-flash",
        "api_url": "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent",
        "api_key": "AIzaSyCoFRhRa1zamcMIqIxIuBChtehsoRE5AUM"
    },
    {
        "name": "Meta Llama 4 Maverick",
        "id": "meta-llama/llama-4-maverick:free",
        "api_url": "https://openrouter.ai/api/v1/chat/completions",
        "api_key": "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",
        "provider": "openrouter"
    }

]
HUMAN_PLAYER_DEFAULT_NAME = "Jugador Humano"

def _crear_entrada_estadisticas_default():
    return {
        "partidas_jugadas": 0,
        "victorias": 0,
        "robos": 0,
        "plantadas": 0,
        "tiempos": [], # Solo para IAs
        "puntos_totales": 0
    }

# Diccionario global para estadísticas. Se poblará dinámicamente.
estadisticas = {}


# Configuración del juego
CARTAS = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']
REPETICIONES_CARTAS = 4


class Player:
    def __init__(self, name):
        self.name = name
        self.is_human = False

    def confirm_rules(self, zona_personal, puntuacion_permanente):
        raise NotImplementedError

    def decide_action(self, carta_robada_del_mazo, zona_temporal, zona_personal, puntuacion_permanente, zonas_personales_oponentes, mazo_restante):
        raise NotImplementedError

    def decide_steal(self, carta_a_robar, victima_nombre, zona_temporal_propia):
        raise NotImplementedError
    
    @property
    def model_name(self): 
        return self.name


class IAClient(Player):
    def __init__(self, name="IA Player"): 
        super().__init__(name)
        self.selected_model = None

    def select_model(self, choice=None, model_config=None):
        if model_config:
            self.selected_model = model_config
            self.name = self.selected_model['name']
        elif choice is not None:
            # Este path es menos común si se usa app.py, pero se mantiene por si acaso
            from rich.prompt import IntPrompt # Import local
            global MODELS 
            self.selected_model = MODELS[choice - 1] 
            self.name = self.selected_model['name']
        else:
            # Fallback interactivo si no se provee ni choice ni model_config
            from rich.prompt import IntPrompt # Import local
            console.print("[bold yellow]Selecciona un modelo de IA (fallback interactivo):[/bold yellow]")
            for i, model_data in enumerate(MODELS, 1):
                console.print(f"{i}. {model_data['name']} ({model_data['id']})")
            choice_num = IntPrompt.ask("Selecciona el número del modelo", default=1)
            self.selected_model = MODELS[choice_num - 1]
            self.name = self.selected_model['name']

    def _send_api_request(self, message_content):
        start_time = time.time()
        headers = {"Content-Type": "application/json"}
        model = self.selected_model
        api_url = model["api_url"]
        model_id = model["id"]
        
        system_message = """Debes responder SOLO con un objeto JSON válido sin texto adicional ni markdown.
        Para decisiones de robar o plantarse:
        - accion: "continuar" o "plantarse"
        Para el mensaje inicial:
        - mensaje: "Entendido"
        """
        max_retries = 2 # Número de reintentos además del intento inicial
        retry_delay_base = 2 # Segundos base para el delay
        request_successful = False
        response_content_str = '{"accion": "plantarse"}' # Default seguro
        
        for attempt in range(max_retries + 1): # Intento inicial + reintentos
            try:
                # console.print(f"[dim]API Req: {model['name']} (Attempt {attempt+1}/{max_retries+1})[/dim]")
                if attempt > 0: # Aplicar delay exponencial para reintentos
                    time.sleep(retry_delay_base * (2 ** (attempt -1)) + random.uniform(0,1))
                else: # Pequeño delay inicial para distribuir cargas si múltiples IAs empiezan al mismo tiempo
                    time.sleep(random.uniform(0.1, 0.5))
                
                current_api_url = api_url
                payload_dict = {}

                if "generativelanguage.googleapis.com" in api_url:
                    if "?" not in current_api_url:
                        current_api_url += f"?key={model['api_key']}"
                    payload_dict = {
                        "contents": [{"parts": [{"text": system_message + "\n\n" + message_content}]}],
                        "generationConfig": {"temperature": 0.7, "maxOutputTokens": 1024}
                    }
                else: # Para OpenRouter y DeepSeek
                    headers["Authorization"] = f"Bearer {model['api_key']}"
                    payload_dict = {
                        "model": model_id,
                        "messages": [
                            {"role": "system", "content": system_message},
                            {"role": "user", "content": message_content}
                        ]
                    }
                
                response = requests.post(current_api_url, headers=headers, json=payload_dict, timeout=60)
                
                if response.status_code in (429, 503): # Rate limit o servicio no disponible
                    if attempt == max_retries: break # No reintentar más en el último intento
                    # console.print(f"[yellow]API {model['name']}: Error {response.status_code}. Reintentando...[/yellow]")
                    continue # Pasa al siguiente intento (con delay manejado al inicio del bucle)
                
                response.raise_for_status() # Lanza HTTPError para otros códigos 4xx/5xx
                
                result = response.json()
                extracted_text = ""
                if "generativelanguage.googleapis.com" in api_url:
                    extracted_text = result.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")
                else: # OpenRouter, DeepSeek
                    extracted_text = result.get("choices", [{}])[0].get("message", {}).get("content", "")
                
                response_content_str = self._clean_json_response(extracted_text)
                request_successful = True
                break # Éxito, salir del bucle de reintentos

            except requests.exceptions.Timeout:
                # console.print(f"[red]API {model['name']}: Timeout (Attempt {attempt+1})[/red]")
                if attempt == max_retries: break
            except requests.exceptions.RequestException as e:
                # console.print(f"[red]API {model['name']}: Request Error (Attempt {attempt+1}): {e}[/red]")
                if attempt == max_retries: break 
            except Exception as e: # Capturar otros errores como JSONDecodeError
                # console.print(f"[red]API {model['name']}: Unexpected Error (Attempt {attempt+1}): {type(e).__name__} - {e}[/red]")
                # response_text_debug = response.text if 'response' in locals() and hasattr(response, 'text') else "No response.text"
                # console.print(f"[dim]Response text for debug: {response_text_debug[:200]}...[/dim]")
                if attempt == max_retries: break
        
        end_time = time.time()
        response_time = end_time - start_time
        if not request_successful:
            console.print(f"[bold red]API FALLO FINAL: {model['name']} después de {max_retries+1} intentos. Usando default: {response_content_str}[/bold red]")
        return response_content_str, response_time, request_successful

    def _clean_json_response(self, response_text):
        response_text = response_text.replace('```json', '').replace('```', '').strip()
        import re
        json_pattern = re.compile(r'(\{.*\})', re.DOTALL)
        match = json_pattern.search(response_text)
        json_str = match.group(1) if match else response_text
        try:
            data = json.loads(json_str)
            if "accion" in data or "mensaje" in data:
                return json.dumps(data)
            # console.print(f"[yellow]JSON válido pero sin 'accion' o 'mensaje': {json_str}[/yellow]")
            return '{"accion": "plantarse"}' 
        except json.JSONDecodeError:
            # console.print(f"[red]Error al decodificar JSON de IA: '{json_str[:100]}...'[/red]")
            return '{"accion": "plantarse"}'

    def confirm_rules(self, zona_personal, puntuacion_permanente):
        prompt = crear_prompt_inicial(zona_personal, puntuacion_permanente)
        response_json_str, _, success = self._send_api_request(prompt)
        if success:
            try:
                return json.loads(response_json_str).get("mensaje") == "Entendido"
            except json.JSONDecodeError: return False
        return False

    def decide_action(self, carta_robada_del_mazo, zona_temporal, zona_personal, puntuacion_permanente, zonas_personales_oponentes, mazo_restante):
        prompt = f"""
Estás jugando a 'Pelusas'. Acabas de robar la carta '{carta_robada_del_mazo}' del mazo.
Tu zona temporal actual (después de añadir '{carta_robada_del_mazo}') es: {zona_temporal}
Tu zona_personal actual es: {zona_personal}
Tu puntuación permanente es: {puntuacion_permanente}
Zonas personales de otros jugadores: {zonas_personales_oponentes}
Cartas restantes en el mazo: {mazo_restante}
Decide si quieres 'continuar' robando o 'plantarse'.
Responde en JSON con: 'accion' ('continuar' o 'plantarse').
"""
        response_json_str, tiempo_respuesta, _ = self._send_api_request(prompt)
        try:
            decision = json.loads(response_json_str).get("accion", "plantarse")
        except json.JSONDecodeError:
            decision = "plantarse"
        return decision, tiempo_respuesta

    def decide_steal(self, carta_a_robar, victima_nombre, zona_temporal_propia):
        # IA siempre intenta robar si se da la condición, no se le pregunta explícitamente.
        return True 


class HumanPlayer(Player):
    def __init__(self, name=HUMAN_PLAYER_DEFAULT_NAME):
        super().__init__(name)
        self.is_human = True
        self.callback_handler = None # Será asignado por app.py

    def confirm_rules(self, zona_personal, puntuacion_permanente):
        if self.callback_handler:
            details = {
                "zona_personal": list(zona_personal), 
                "puntuacion_permanente": puntuacion_permanente
            }
            response = self.callback_handler.request_human_action("confirm_rules", details)
            if response is None: 
                console.log(f"[HumanPlayer:{self.name}] No se recibió confirmación de reglas (timeout/error). Asumiendo 'No'.")
                return False
            return response 
        else: # Fallback a consola
            from rich.prompt import Confirm as RichConfirm 
            console.print(f"\n[bold magenta]{self.name}[/bold magenta], (Fallback Consola) Revisa las reglas.")
            return RichConfirm.ask("¿Entiendes las reglas y estás listo?", default=True)

    def decide_action(self, carta_robada_del_mazo, zona_temporal, zona_personal, puntuacion_permanente, zonas_personales_oponentes, mazo_restante):
        if self.callback_handler:
            details = {
                # La 'carta_robada_del_mazo' es la que disparó esta decisión.
                # 'zona_temporal' ya incluye esta carta.
                "carta_robada_del_mazo": carta_robada_del_mazo, 
                "zona_temporal": list(zona_temporal), 
                "zona_personal": list(zona_personal),
                "puntuacion_permanente": puntuacion_permanente,
                "zonas_personales_oponentes": [list(zp) for zp in zonas_personales_oponentes],
                "mazo_restante": mazo_restante
            }
            response = self.callback_handler.request_human_action("decide_action", details)
            if response: # response es una tupla (decision_str, tiempo_float)
                return response
            else: # Timeout o error
                console.log(f"[HumanPlayer:{self.name}] No se recibió decisión de acción (timeout/error). Default a 'plantarse'.")
                return "plantarse", 0.0
        else: # Fallback a consola
            from rich.prompt import Prompt as RichPrompt 
            console.print(f"\n[bold magenta]{self.name}[/bold magenta], (Fallback Consola) tu zona temporal es: [bold]{list(zona_temporal)}[/bold]")
            decision = RichPrompt.ask("¿Qué quieres hacer?", choices=["continuar", "plantarse"], default="continuar")
            return decision, 0.0

    def decide_steal(self, carta_a_robar, victima_nombre, zona_temporal_propia):
        if self.callback_handler:
            details = {
                "carta_a_robar": carta_a_robar, 
                "victima_nombre": victima_nombre,
                "zona_temporal_propia": list(zona_temporal_propia) 
            }
            response = self.callback_handler.request_human_action("decide_steal", details)
            if response is None: 
                console.log(f"[HumanPlayer:{self.name}] No se recibió decisión de robo (timeout/error). Default a 'False' (no robar).")
                return False
            return response 
        else: # Fallback a consola
            from rich.prompt import Confirm as RichConfirm
            console.print(f"\n[bold magenta]{self.name}[/bold magenta], (Fallback Consola) puedes robar [bold]{carta_a_robar}[/bold] de {victima_nombre}.")
            return RichConfirm.ask(f"¿Quieres robar la carta [bold]{carta_a_robar}[/bold] de [bold]{victima_nombre}[/bold]?", default=True)


def test_models():
    operative_models_configs = []
    # Nota: Los eventos de SocketIO para testeo se manejan en app.py si se llama desde allí.
    # Esta función es más para testeo directo desde pelusas.py.
    
    for model_config in MODELS:
        console.print(f"[yellow]Testeando {model_config['name']}...[/yellow]")
        test_success = False
        try:
            ia_test = IAClient()
            ia_test.select_model(model_config=model_config)
            
            test_message_content = "Hola, ¿estás listo para jugar? Responde con {\"mensaje\": \"Entendido\"}"
            response_json_str, _, success_flag = ia_test._send_api_request(test_message_content)
            
            if success_flag:
                try:
                    if json.loads(response_json_str).get("mensaje") == "Entendido":
                        operative_models_configs.append(model_config)
                        console.print(f"[green]{model_config['name']} está operativo.[/green]")
                        test_success = True
                    else:
                        console.print(f"[red]{model_config['name']} no respondió 'Entendido'. Respuesta: {response_json_str}[/red]")
                except json.JSONDecodeError:
                     console.print(f"[red]{model_config['name']} devolvió JSON inválido: {response_json_str}[/red]")
            else:
                console.print(f"[red]{model_config['name']} falló la solicitud API.[/red]")
                
        except Exception as e:
            console.print(f"[red]Excepción al testear {model_config['name']}: {type(e).__name__} - {e}[/red]")
        
        time.sleep(1) # Pausa entre tests
    
    if not operative_models_configs:
        console.print("[bold red]¡Advertencia! No se encontraron modelos operativos durante el testeo.[/bold red]")
    return operative_models_configs


def crear_mazo():
    mazo = CARTAS * REPETICIONES_CARTAS
    random.shuffle(mazo)
    return mazo

def valor_carta(carta):
    return int(carta)

def crear_prompt_inicial(zona_personal, puntuacion_permanente):
    return f"""
Estás jugando a 'Pelusas' con otras IAs. Aquí están las reglas completas:
- En cada turno, puedes robar cartas del mazo una a una y colocarlas en tu zona temporal.
- Si robas una carta del mazo que hace que tengas dos cartas iguales en tu zona temporal, ¡pierdes todas las cartas de tu zona temporal! Se descartan, y tu turno termina.
- Puedes plantarte en cualquier momento. Si lo haces, las cartas de tu zona temporal pasan a tu zona personal.
- Las cartas en tu zona personal NO son puntos definitivos hasta que te vuelva a tocar jugar. Si sobreviven ese ciclo, se vuelven puntuación permanente.
- Si en tu turno sacas una carta del mazo que coincide con una de tu zona personal (y tienes 3 o más cartas en dicha zona personal), pierdes toda esa zona personal y tu turno termina. Tu zona temporal también se pierde.
- Si robas una carta del mazo, y esta carta coincide con una que un oponente tiene en su zona personal, puedes decidir robar esa carta del oponente. Si lo haces, esa carta se añade a tu zona temporal. Si esto causa un duplicado en tu zona temporal, pierdes tu zona temporal y tu turno termina.
- La partida termina cuando se acaba el mazo. Solo cuentan las cartas que hayan pasado a puntuación permanente.
- Gana quien sume más puntos con sus cartas definitivas.
Tu zona personal actual es: {zona_personal}
Tu puntuación permanente hasta ahora es: {puntuacion_permanente}
Este es solo el inicio. Pronto empezará el juego.
Responde en JSON: {{"mensaje": "Entendido"}}
"""

def log_to_excel(partida_num, data):
    filename = "juego_pelusas_log.xlsx"
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"]) # Eliminar hoja por defecto si es nueva

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sheet_name_base = f"Partida_{partida_num}"
    sheet_name = f"{sheet_name_base}_{timestamp}"
    sheet_name = sheet_name[:31] # Límite de Excel para nombres de hoja
    
    # Manejar duplicados de nombre de hoja
    original_sheet_name = sheet_name
    count = 1
    while sheet_name in wb.sheetnames:
        suffix = f"_{count}"
        # Asegurar que el nombre base + sufijo no exceda el límite
        sheet_name = f"{original_sheet_name[:31-len(suffix)]}{suffix}"
        count += 1

    sheet = wb.create_sheet(sheet_name)
    headers = ["Turno", "Jugador", "Acción", "Carta Robada Mazo", "Zona Temporal", "Zona Personal", "Puntuación Permanente", "Tiempo Respuesta (s)", "Mazo Restante"]
    for col, header in enumerate(headers, 1):
        sheet[f"{get_column_letter(col)}1"] = header
    
    row_idx = 2
    for entry in data:
        sheet[f"A{row_idx}"] = entry.get("turno", "")
        sheet[f"B{row_idx}"] = entry.get("jugador", "")
        sheet[f"C{row_idx}"] = entry.get("accion", "")
        sheet[f"D{row_idx}"] = entry.get("carta_robada_mazo", "")
        sheet[f"E{row_idx}"] = str(entry.get("zona_temporal", []))
        sheet[f"F{row_idx}"] = str(entry.get("zona_personal", []))
        sheet[f"G{row_idx}"] = entry.get("puntuacion_permanente", 0)
        sheet[f"H{row_idx}"] = f"{entry.get('tiempo_respuesta', 0):.2f}" if isinstance(entry.get('tiempo_respuesta'), float) else entry.get('tiempo_respuesta', 0)
        sheet[f"I{row_idx}"] = entry.get("mazo_restante", 0)
        row_idx += 1
    try:
        wb.save(filename)
    except Exception as e:
        console.print(f"[red]Error guardando Excel '{filename}': {e}. Puede estar abierto por otro programa.[/red]")


def seleccionar_jugadores(num_jugadores_total=4, operative_model_configs=None, modo_seleccion_ia='random', num_human_players=0):
    if operative_model_configs is None:
        operative_model_configs = list(MODELS) # Usar una copia de la lista global por defecto
    
    jugadores = []
    num_ia_players = num_jugadores_total - num_human_players

    ia_configs_seleccionadas = []
    if num_ia_players > 0:
        if not operative_model_configs:
            console.print("[red]No hay modelos de IA operativos para seleccionar IA players.[/red]")
        else:
            if modo_seleccion_ia == 'random':
                num_to_sample = min(num_ia_players, len(operative_model_configs))
                ia_configs_seleccionadas = random.sample(operative_model_configs, num_to_sample)
            else: # 'static' o default: tomar los primeros N
                ia_configs_seleccionadas = operative_model_configs[:min(num_ia_players, len(operative_model_configs))]

    for config in ia_configs_seleccionadas:
        ia_player = IAClient()
        ia_player.select_model(model_config=config)
        jugadores.append(ia_player)

    for i in range(num_human_players):
        human_name = f"{HUMAN_PLAYER_DEFAULT_NAME} {i+1}" if num_human_players > 1 else HUMAN_PLAYER_DEFAULT_NAME
        jugadores.append(HumanPlayer(name=human_name))

    random.shuffle(jugadores) # Barajar el orden de todos los jugadores
    
    if len(jugadores) != num_jugadores_total: # Esto podría pasar si no hay suficientes IAs operativas
         console.print(f"[yellow]Advertencia: Se configuraron {len(jugadores)} jugadores en lugar de los {num_jugadores_total} solicitados.[/yellow]")

    return jugadores


def generar_excel_resumido():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Resumen_Partidas"
    headers = ["Jugador", "Partidas Jugadas", "Victorias", "% Victorias", "Tiempo Medio Respuesta (IA)", "Robos Promedio", "Plantadas Promedio", "Puntos Promedio"]
    for col, header in enumerate(headers, 1):
        sheet[f"{get_column_letter(col)}1"] = header
    
    row = 2
    for player_name, data_raw in estadisticas.items():
        # Asegurar que la entrada de datos existe y está completa
        data = data_raw
        if not isinstance(data, dict): data = _crear_entrada_estadisticas_default()
        else: # Asegurar que todas las claves existen
            default_entry = _crear_entrada_estadisticas_default()
            for key, val_default in default_entry.items():
                data.setdefault(key, val_default)

        partidas_jugadas = data["partidas_jugadas"]
        if partidas_jugadas == 0: continue # No mostrar jugadores que no jugaron
        
        victorias = data["victorias"]
        porcentaje_victorias = (victorias / partidas_jugadas) * 100 if partidas_jugadas > 0 else 0
        
        tiempos_respuesta_ia = data.get("tiempos", []) # Tiempos solo aplica a IAs
        tiempo_medio_ia = sum(tiempos_respuesta_ia) / len(tiempos_respuesta_ia) if tiempos_respuesta_ia else 0
        
        robos_promedio = data["robos"] / partidas_jugadas if partidas_jugadas > 0 else 0
        plantadas_promedio = data["plantadas"] / partidas_jugadas if partidas_jugadas > 0 else 0
        puntos_promedio = data["puntos_totales"] / partidas_jugadas if partidas_jugadas > 0 else 0
        
        sheet[f"A{row}"] = player_name
        sheet[f"B{row}"] = partidas_jugadas
        sheet[f"C{row}"] = victorias
        sheet[f"D{row}"] = f"{porcentaje_victorias:.2f}%"
        sheet[f"E{row}"] = f"{tiempo_medio_ia:.2f}" if tiempo_medio_ia > 0 else "N/A"
        sheet[f"F{row}"] = f"{robos_promedio:.2f}"
        sheet[f"G{row}"] = f"{plantadas_promedio:.2f}"
        sheet[f"H{row}"] = f"{puntos_promedio:.2f}"
        row += 1
    
    try:
        wb.save("resumen_partidas_pelusas.xlsx")
        console.print("[green]Excel resumen generado: resumen_partidas_pelusas.xlsx[/green]")
    except Exception as e:
        console.print(f"[red]Error guardando Excel resumen: {e}. Puede estar abierto por otro programa.[/red]")


class GameObserver: 
    def on_game_start(self, jugadores_lista): pass
    def on_turn_start(self, jugador_obj): pass
    def on_card_drawn(self, jugador_obj, carta, zonas_temporales, zonas_personales, puntuaciones_permanentes, mazo): pass
    def on_decision(self, jugador_obj, decision_str): pass
    def on_turn_end(self, jugador_obj): pass
    def on_game_end(self, ganador_obj, puntuaciones_permanentes_lista): pass
    def on_temporal_zone_lost(self, jugador_obj, causa="desconocida"): pass 
    def on_personal_zone_lost(self, jugador_obj, causa="desconocida"): pass 
    def on_steal_from_player(self, jugador_ladron_obj, jugador_victima_obj, carta_robada_str): pass
    def update_game_state(self, jugadores_lista, zonas_temporales, zonas_personales, puntuaciones_permanentes_lista, mazo): pass


def main(jugadores_lista: list[Player], partida_num: int, observer: GameObserver = None):
    if observer is None:
        observer = GameObserver() 

    num_jugadores = len(jugadores_lista)
    if num_jugadores == 0:
        console.print("[red]No hay jugadores para iniciar la partida.[/red]")
        return

    # Inicializar/Actualizar estadísticas para todos los jugadores
    for jugador_obj in jugadores_lista:
        if jugador_obj.name not in estadisticas:
            estadisticas[jugador_obj.name] = _crear_entrada_estadisticas_default()
        estadisticas[jugador_obj.name]["partidas_jugadas"] += 1
    
    console.print(f"[cyan]Iniciando partida {partida_num} con {num_jugadores} jugadores: {[j.name for j in jugadores_lista]}[/cyan]")
    mazo = crear_mazo()
    zonas_temporales = [[] for _ in range(num_jugadores)] # Lista de listas, una por jugador
    zonas_personales = [[] for _ in range(num_jugadores)] # Lista de listas
    puntuaciones_permanentes = [0 for _ in range(num_jugadores)] # Lista de ints
    turno_idx = 0 
    log_data = []

    observer.on_game_start(list(jugadores_lista)) # Enviar copia de la lista de jugadores

    # Confirmación de reglas
    for i, jugador_obj in enumerate(jugadores_lista):
        # HumanPlayer.confirm_rules usará el callback_handler si está asignado
        rules_confirmed = jugador_obj.confirm_rules(zonas_personales[i], puntuaciones_permanentes[i])
        if not rules_confirmed:
            console.print(f"[red]Error: {jugador_obj.name} no confirmó las reglas. Abortando partida {partida_num}.[/red]")
            # Se podría decrementar "partidas_jugadas" aquí o manejarlo como partida inválida
            return

    try:
        partida_activa = True
        while len(mazo) > 0 and partida_activa:
            jugador_actual_obj = jugadores_lista[turno_idx]
            
            # Referencias a las zonas del jugador actual para facilitar lectura
            zona_temporal_actual = zonas_temporales[turno_idx] # Es una referencia a la lista del jugador actual
            zona_personal_actual = zonas_personales[turno_idx] # Es una referencia

            observer.on_turn_start(jugador_actual_obj)

            # Convertir zona personal en puntos permanentes si sobrevivió la ronda
            if zona_personal_actual: # Si hay cartas en zona personal del turno anterior
                puntos_ganados = sum(valor_carta(c) for c in zona_personal_actual)
                puntuaciones_permanentes[turno_idx] += puntos_ganados
                console.print(f"[green]{jugador_actual_obj.name} convierte {list(zona_personal_actual)} en {puntos_ganados} puntos permanentes.[/green]")
                zona_personal_actual.clear() # Limpiar la zona personal del jugador actual
                observer.update_game_state(list(jugadores_lista), zonas_temporales, zonas_personales, puntuaciones_permanentes, mazo)

            console.print(f"\n[bold cyan]Turno de {jugador_actual_obj.name}[/bold cyan]")
            log_data.append({
                "turno": len(log_data) + 1, "jugador": jugador_actual_obj.name, "accion": "inicio_turno",
                "zona_temporal": list(zona_temporal_actual), "zona_personal": list(zona_personal_actual), # Copias para el log
                "puntuacion_permanente": puntuaciones_permanentes[turno_idx], "mazo_restante": len(mazo)
            })

            turno_del_jugador_continua = True
            while len(mazo) > 0 and turno_del_jugador_continua:
                if not partida_activa: break # Si el mazo se acabó en un turno anterior y se marcó fin de partida

                # --- INICIO DE UN ROBO DE CARTA DEL MAZO ---
                carta_robada_del_mazo = mazo.pop(0)
                console.print(f"[blue]{jugador_actual_obj.name} robó una carta del mazo: {carta_robada_del_mazo}[/blue]")
                
                # 1. Verificar PRIMERO si la carta robada causaría un duplicado en zona temporal
                # PERO solo aplicar la regla si ya tienes 3 o más cartas en zona temporal
                if len(zona_temporal_actual) >= 3 and carta_robada_del_mazo in zona_temporal_actual:
                    console.print(f"[red]{jugador_actual_obj.name} robó {carta_robada_del_mazo} que ya está en su zona temporal ({list(zona_temporal_actual)}).[/red]")
                    console.print(f"[red]Ya tiene 3+ cartas en zona temporal, por lo que pierde su zona temporal. Turno termina.[/red]")
                    
                    observer.on_temporal_zone_lost(jugador_actual_obj, causa="duplicado_mazo")
                    zona_temporal_actual.clear()
                    
                    observer.update_game_state(list(jugadores_lista), zonas_temporales, zonas_personales, puntuaciones_permanentes, mazo)
                    turno_del_jugador_continua = False # Termina el turno del jugador
                    break # Salir del bucle de robar cartas para este jugador
                
                # 2. Añadir carta a la zona temporal (solo si no hay duplicado problemático)
                zona_temporal_actual.append(carta_robada_del_mazo)
                console.print(f"[dim]{jugador_actual_obj.name} - Zona Temporal ahora: {list(zona_temporal_actual)}[/dim]")

                # Notificar al observer y loguear *después* de añadirla a la zona temporal
                observer.on_card_drawn(jugador_actual_obj, carta_robada_del_mazo, zonas_temporales, zonas_personales, puntuaciones_permanentes, mazo)
                log_data.append({
                    "turno": len(log_data) + 1, "jugador": jugador_actual_obj.name, "accion": "robar_del_mazo",
                    "carta_robada_mazo": carta_robada_del_mazo, "zona_temporal": list(zona_temporal_actual),
                    "zona_personal": list(zona_personal_actual), "puntuacion_permanente": puntuaciones_permanentes[turno_idx],
                    "mazo_restante": len(mazo)
                })

                # 3. Verificar Regla de Avaricia (Pérdida de Zona Personal)
                # Ocurre si la carta robada del MAZO coincide con una en ZONA PERSONAL (y hay 3+ allí)
                if carta_robada_del_mazo in zona_personal_actual and len(zona_personal_actual) >= 3:
                    console.print(f"[red]{jugador_actual_obj.name} (AVARICIA) robó {carta_robada_del_mazo} que está en su zona personal ({list(zona_personal_actual)}) que tiene {len(zona_personal_actual)} cartas.[/red]")
                    console.print(f"[red]Pierde TODA su zona personal ({list(zona_personal_actual)}) y su zona temporal ({list(zona_temporal_actual)}). Turno termina.[/red]")
                    
                    observer.on_personal_zone_lost(jugador_actual_obj, causa="avaricia")
                    zona_personal_actual.clear()
                    observer.on_temporal_zone_lost(jugador_actual_obj, causa="avaricia_termina_turno")
                    zona_temporal_actual.clear()
                    
                    observer.update_game_state(list(jugadores_lista), zonas_temporales, zonas_personales, puntuaciones_permanentes, mazo)
                    turno_del_jugador_continua = False # Termina el turno del jugador
                    break # Salir del bucle de robar cartas para este jugador

                # 4. Oportunidad de Robar de Oponentes
                # Solo si el turno no ha terminado por las reglas anteriores.
                # La carta que se usa para comparar con las zonas personales de oponentes es la `carta_robada_del_mazo`.
                carta_para_comparar_con_oponentes = carta_robada_del_mazo 
                for i_otro_jugador in range(num_jugadores):
                    if i_otro_jugador == turno_idx: continue # No robarse a sí mismo

                    if carta_para_comparar_con_oponentes in zonas_personales[i_otro_jugador]:
                        console.print(f"[magenta]{jugador_actual_obj.name} puede robar '{carta_para_comparar_con_oponentes}' de la zona personal de {jugadores_lista[i_otro_jugador].name} ({list(zonas_personales[i_otro_jugador])}).[/magenta]")
                        
                        # HumanPlayer.decide_steal usará el callback_handler
                        quiere_robar_de_oponente = jugador_actual_obj.decide_steal(
                            carta_para_comparar_con_oponentes, 
                            jugadores_lista[i_otro_jugador].name, 
                            list(zona_temporal_actual) # Zona temporal actual ANTES de añadir la carta del oponente
                        )

                        if quiere_robar_de_oponente:
                            console.print(f"[green]{jugador_actual_obj.name} decide robar '{carta_para_comparar_con_oponentes}' de {jugadores_lista[i_otro_jugador].name}.[/green]")
                            zonas_personales[i_otro_jugador].remove(carta_para_comparar_con_oponentes) # Quitar de la víctima
                            zona_temporal_actual.append(carta_para_comparar_con_oponentes) # Añadir a la zona temporal del ladrón
                            console.print(f"[dim]{jugador_actual_obj.name} - Zona Temporal tras robo a oponente: {list(zona_temporal_actual)}[/dim]")
                            
                            observer.on_steal_from_player(jugador_actual_obj, jugadores_lista[i_otro_jugador], carta_para_comparar_con_oponentes)
                            estadisticas[jugador_actual_obj.name].setdefault("robos",0) # Asegurar que la clave existe
                            estadisticas[jugador_actual_obj.name]["robos"] += 1
                            
                            # Al robar de oponentes, NUNCA se pierde la zona temporal por duplicados
                            # Esto es parte de las reglas del juego
                        else:
                            console.print(f"[yellow]{jugador_actual_obj.name} decide NO robar '{carta_para_comparar_con_oponentes}' de {jugadores_lista[i_otro_jugador].name}.[/yellow]")

                # 5. Decidir Continuar o Plantarse (si el turno aún no ha terminado por las reglas anteriores)
                # HumanPlayer.decide_action usará el callback_handler
                decision_str, tiempo_resp = jugador_actual_obj.decide_action(
                    carta_robada_del_mazo, # La carta que se robó del mazo en este paso
                    list(zona_temporal_actual), # Zona temporal actual
                    list(zona_personal_actual),
                    puntuaciones_permanentes[turno_idx],
                    [zonas_personales[j] for j in range(num_jugadores) if j != turno_idx], # Zonas personales de otros
                    len(mazo)
                )

                if not jugador_actual_obj.is_human: # Solo registrar tiempo para IAs
                    estadisticas[jugador_actual_obj.name].setdefault("tiempos", []).append(tiempo_resp)
                
                observer.on_decision(jugador_actual_obj, decision_str)
                log_data.append({
                    "turno": len(log_data) + 1, "jugador": jugador_actual_obj.name, "accion": f"decision_{decision_str}",
                    "carta_robada_mazo": carta_robada_del_mazo, "zona_temporal": list(zona_temporal_actual),
                    "zona_personal": list(zona_personal_actual), "puntuacion_permanente": puntuaciones_permanentes[turno_idx],
                    "tiempo_respuesta": tiempo_resp if not jugador_actual_obj.is_human else 0.0,
                    "mazo_restante": len(mazo)
                })

                if decision_str == "plantarse":
                    console.print(f"[yellow]{jugador_actual_obj.name} decide plantarse. Zona temporal ({list(zona_temporal_actual)}) pasa a personal.[/yellow]")
                    zonas_personales[turno_idx].extend(zona_temporal_actual) # Mover de temporal a personal del jugador actual
                    zona_temporal_actual.clear()
                    estadisticas[jugador_actual_obj.name].setdefault("plantadas", 0) # Asegurar clave
                    estadisticas[jugador_actual_obj.name]["plantadas"] += 1
                    observer.update_game_state(list(jugadores_lista), zonas_temporales, zonas_personales, puntuaciones_permanentes, mazo)
                    turno_del_jugador_continua = False # Termina el turno del jugador
                else: # Continuar
                    console.print(f"[yellow]{jugador_actual_obj.name} decide continuar robando.[/yellow]")
                    # El bucle de turno del jugador (`while len(mazo) > 0 and turno_del_jugador_continua:`) continúa
            
            # --- FIN DEL TURNO DEL JUGADOR ACTUAL ---
            observer.on_turn_end(jugador_actual_obj)
            turno_idx = (turno_idx + 1) % num_jugadores # Pasar al siguiente jugador
            if len(mazo) == 0: # Si el mazo se acabó durante el turno de este jugador
                partida_activa = False # Terminar la partida

        # --- FIN DE LA PARTIDA (MAZO AGOTADO) ---
        console.print("[bold green]El mazo se ha acabado. Fin de la partida.[/bold green]")
        
        # Puntuación final: las cartas en zona temporal y personal se convierten en puntos
        for i in range(num_jugadores):
            jugador_obj = jugadores_lista[i]
            # Sumar puntos de zona temporal si quedó algo (ej. mazo se acabó en el turno de otro)
            if zonas_temporales[i]:
                puntos_temp = sum(valor_carta(c) for c in zonas_temporales[i])
                puntuaciones_permanentes[i] += puntos_temp
                console.print(f"[green]{jugador_obj.name} suma {puntos_temp} puntos de su zona temporal ({list(zonas_temporales[i])}) final.[/green]")
                zonas_temporales[i].clear()
            # Sumar puntos de zona personal si quedó algo (ej. se plantó justo antes de acabar el mazo o no tuvo otro turno)
            if zonas_personales[i]: 
                puntos_pers = sum(valor_carta(c) for c in zonas_personales[i])
                puntuaciones_permanentes[i] += puntos_pers
                console.print(f"[green]{jugador_obj.name} suma {puntos_pers} puntos de su zona personal ({list(zonas_personales[i])}) final.[/green]")
                zonas_personales[i].clear()
        
        observer.update_game_state(list(jugadores_lista), zonas_temporales, zonas_personales, puntuaciones_permanentes, mazo)
        
        # Determinar ganador
        idx_ganador = -1
        max_puntos = -1
        # Simple: el de mayor puntuación, desempate por orden si es necesario
        for i in range(num_jugadores):
            if puntuaciones_permanentes[i] > max_puntos:
                max_puntos = puntuaciones_permanentes[i]
                idx_ganador = i
            elif puntuaciones_permanentes[i] == max_puntos:
                # En caso de empate, el de menor índice (primero en la lista original de jugadores) gana.
                # Podrías implementar otra lógica de desempate aquí si lo deseas.
                pass 
        
        ganador_obj = jugadores_lista[idx_ganador] if idx_ganador != -1 else None
        
        if ganador_obj:
            console.print(f"\n🏆 [bold green]¡{ganador_obj.name} ha ganado con {puntuaciones_permanentes[idx_ganador]} puntos permanentes![/bold green]")
            estadisticas[ganador_obj.name].setdefault("victorias", 0) # Asegurar clave
            estadisticas[ganador_obj.name]["victorias"] += 1
        else:
            console.print("\n[bold yellow]No hubo ganador claro o no hay jugadores.[/bold yellow]")
            
        observer.on_game_end(ganador_obj, list(puntuaciones_permanentes)) # Enviar copia de la lista

        # Registrar puntos totales para todos
        for i in range(num_jugadores):
            jugador_obj = jugadores_lista[i]
            console.print(f"{jugador_obj.name}: {puntuaciones_permanentes[i]} puntos permanentes")
            estadisticas[jugador_obj.name].setdefault("puntos_totales", 0) # Asegurar clave
            estadisticas[jugador_obj.name]["puntos_totales"] += puntuaciones_permanentes[i]
        
    except Exception as e:
        console.print(f"[bold red]Error durante la partida {partida_num}: {type(e).__name__} - {e}[/bold red]")
        console.print_exception(show_locals=True) # Mostrar traceback completo y variables locales

    # Guardar log de la partida
    try:
        log_to_excel(partida_num, log_data)
        console.print(f"[green]Log de la partida {partida_num} guardado.[/green]")
    except Exception as e:
        console.print(f"[red]Error al guardar log para partida {partida_num}: {e}[/red]")


if __name__ == "__main__":
    # Imports específicos para ejecución directa en consola
    from rich.prompt import IntPrompt as RichIntPrompt, Confirm as RichConfirm
    
    console.print("[cyan]Modo de ejecución directa de pelusas.py (consola)[/cyan]")
    
    console.print("[cyan]Testeando modelos de IA...[/cyan]")
    operative_model_configs = test_models()

    if not operative_model_configs:
        console.print("[yellow]No hay modelos de IA operativos. Se jugará solo con humanos si se configura así.[/yellow]")
    
    num_total_jugadores_str = RichIntPrompt.ask("Número total de jugadores (1-4)", choices=["1","2","3","4"], default="2")
    num_humanos_str = RichIntPrompt.ask(f"Número de jugadores humanos (0-{num_total_jugadores_str})", default="1", choices=[str(i) for i in range(int(num_total_jugadores_str) + 1)])
    
    num_humanos = int(num_humanos_str)
    num_total_jugadores = int(num_total_jugadores_str)

    if num_humanos > num_total_jugadores:
        console.print("[red]Número de humanos no puede exceder el total de jugadores. Ajustando.[/red]")
        num_humanos = num_total_jugadores

    # Verificar si se necesitan IAs y si hay modelos operativos
    if (num_total_jugadores - num_humanos) > 0 and not operative_model_configs:
        console.print("[red]Se requieren IAs pero no hay modelos operativos. No se puede iniciar la partida como está configurada.[/red]")
        exit()

    jugadores_partida = seleccionar_jugadores(
        num_jugadores_total=num_total_jugadores,
        operative_model_configs=operative_model_configs,
        modo_seleccion_ia='random', # Para consola, 'random' es lo más simple
        num_human_players=num_humanos
    )

    if not jugadores_partida or len(jugadores_partida) != num_total_jugadores :
        console.print("[red]No se pudieron configurar suficientes jugadores para la partida. Saliendo.[/red]")
        exit()

    num_partidas_a_jugar = 1
    if num_humanos == 0: # Si no hay humanos, preguntar si quieren múltiples partidas
        if RichConfirm.ask("¿Jugar una serie de 100 partidas (solo IAs)?", default=False):
            num_partidas_a_jugar = 100
    
    for i in range(1, num_partidas_a_jugar + 1):
        # Para múltiples partidas de IAs, podríamos querer re-seleccionar IAs cada vez para variedad
        if num_humanos == 0 and num_partidas_a_jugar > 1:
             jugadores_ronda_i = seleccionar_jugadores(
                num_jugadores_total=num_total_jugadores,
                operative_model_configs=operative_model_configs, # Usar los operativos testeados al inicio
                modo_seleccion_ia='random',
                num_human_players=0
            )
        else:
            # Para la primera partida o si es una sola partida (o con humanos), usar la selección inicial.
            jugadores_ronda_i = list(jugadores_partida) # Crear una copia para evitar modificar la original si se reasigna

        if jugadores_ronda_i and len(jugadores_ronda_i) == num_total_jugadores :
            main(jugadores_ronda_i, partida_num=i)
        else:
            console.print(f"[red]No se pudieron configurar jugadores para la partida {i}. Se omitirá esta ronda.[/red]")
            # Si `main` modifica la lista de jugadores (no debería), esto podría ser un problema.
            # Por eso se pasa `list(jugadores_partida)` para pasar una copia.
            if not jugadores_ronda_i : break # Salir si no hay jugadores para la ronda
        
        if num_partidas_a_jugar > 1 and i < num_partidas_a_jugar:
            console.print(f"[magenta]Siguiente partida ({i+1}/{num_partidas_a_jugar}) en 3 segundos...[/magenta]")
            time.sleep(3)


    # Generar resumen si se jugaron partidas y hay datos en `estadisticas`
    if any(data.get("partidas_jugadas", 0) > 0 for data in estadisticas.values()):
         generar_excel_resumido()
    else:
        console.print("[yellow]No se jugaron partidas o no hay datos, no se generará resumen.[/yellow]")